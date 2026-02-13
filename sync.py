#!/usr/bin/env python3
"""
Sync-Script: Artikeldaten aus dem Shopware Produktfeed aktualisieren.

Nutzung:
    python3 sync.py

Merged zwei Datenquellen:
  - Shopware Produktfeed → aktuelle Produkte, Kategorien, Preise
  - data.xlsx → offizielle Artikelbezeichnungen (Art_Bezeichnung)

Falls data.xlsx nicht vorhanden ist, wird nur der Feed verwendet.
"""

import subprocess
import xml.etree.ElementTree as ET
import json
import os
from datetime import date

FEED_URL = "https://www.jeanlen.de/store-api/product-export/SWPERW1QSK80Z1DSAHHWCDIZRA/chatchamp.xml"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FEED_FILE = os.path.join(SCRIPT_DIR, "feed.xml")
EXCEL_FILE = os.path.join(SCRIPT_DIR, "data.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "articles.json")

# Kategorie-Keyword-Mapping für die Chatsuche
CATEGORY_KEYWORDS = {
    "Haare": ["haare", "haar", "haarpflege", "shampoo", "conditioner", "spülung", "haarkur", "haarmaske", "leave-in", "haarspray", "locken", "volumen", "haarwäsche", "kopfhaut"],
    "Pflege": ["pflege", "körperpflege", "hautpflege"],
    "Pflege > Körper": ["körper", "body lotion", "bodylotion", "körperlotion", "body butter", "duschgel", "shower gel", "dusche", "peeling", "body oil", "körperöl", "badezusatz", "bath", "body wash", "schaumbad"],
    "Pflege > Hände": ["handpflege", "handcreme", "hand cream", "handbalsam", "hand balm", "hände"],
    "Pflege > Handseifen": ["handseife", "hand soap", "seife", "schaumseife", "handseifen"],
    "Pflege > Gesicht & Lippen": ["gesichtspflege", "gesicht", "face", "gesichtscreme", "face cream", "serum", "cleanser", "toner", "tagescreme", "nachtcreme", "day cream", "night cream", "reinigungsgel", "augenpflege", "eye cream"],
    "Pflege > Lippen": ["lip balm", "lip treatment", "lippenpflege", "lippen", "lip butter", "lippenbalsam", "lippenbalm"],
    "Pflege > Sonne": ["sonne", "sonnencreme", "sonnenmilch", "sunscreen", "spf", "lsf", "after sun", "sonnenschutz", "sonnenpflege", "aloe vera"],
    "Pflege > Deos": ["deo", "deodorant", "antitranspirant", "deos"],
    "Pflege > Kinder & Babys": ["kinder", "baby", "babys", "kids", "kind", "piraten", "feen", "sensibelchen"],
    "Pflege > Sets": ["set", "geschenkset", "giftbox", "geschenkbox", "bundle", "adventskalender"],
    "Home": ["home", "zuhause", "wohnung"],
    "Home > Raumduft & Kerzen": ["raumduft", "duftkerze", "kerze", "candle", "diffusor", "diffuser", "room diffuser", "aromatic", "atmosphere", "wäscheduft"],
    "Home > Handtücher": ["handtuch", "handtücher", "towel"],
    "Home > Reiniger": ["reiniger", "allzweckreiniger", "putzstein", "waschmittel", "weichspüler"],
    "Duft": ["duft", "parfum", "parfüm", "eau de toilette", "edt", "fragrance", "scent"],
    "Haare > Shampoo": ["shampoo", "haarshampoo", "haarwäsche"],
    "Haare > Conditioner": ["conditioner", "spülung", "haarspülung"],
    "Haare > Extrapflege": ["haarkur", "haarmaske", "leave-in", "haarspray", "extrapflege", "aufbauspray", "haaröl", "hair oil", "heat protection", "hitzeschutz"],
    "Sonstiges": ["sonstiges"],
}


def load_excel_names():
    """Offizielle Artikelbezeichnungen aus data.xlsx laden."""
    if not os.path.exists(EXCEL_FILE):
        print("Hinweis: data.xlsx nicht gefunden – verwende nur Feed-Titel.")
        return {}

    try:
        import openpyxl
    except ImportError:
        print("Hinweis: openpyxl nicht installiert – verwende nur Feed-Titel.")
        print("  Installieren mit: pip3 install openpyxl")
        return {}

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Export"]
    names = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        sku = str(row[0]).strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        if sku and name:
            names[sku] = name
    print(f"Excel geladen: {len(names)} offizielle Bezeichnungen")
    return names


def download_feed():
    """Feed per curl herunterladen (umgeht SSL-Probleme in Python)."""
    print(f"Lade Feed von {FEED_URL} ...")
    result = subprocess.run(
        ["curl", "-s", "-o", FEED_FILE, FEED_URL],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Fehler beim Download: {result.stderr}")
    print("Feed heruntergeladen.")


def parse_feed(excel_names):
    """Feed-XML parsen und mit Excel-Namen mergen."""
    root = ET.parse(FEED_FILE).getroot()
    articles = []
    matched = 0
    fallback = 0

    for product in root.findall("product"):
        sku = (product.findtext("ordernumber") or "").strip()
        feed_title = (product.findtext("title") or "").strip()
        feed_title = feed_title.replace(" | JeanLen", "").replace(" | Jean&Len", "").strip()
        availability = (product.findtext("availability") or "").strip()
        price = (product.findtext("price") or "").strip()
        sale_price = (product.findtext("sale_price") or "").strip()
        link = (product.findtext("link") or "").strip()
        gtin = (product.findtext("gtin") or "").strip()
        cat0 = (product.findtext("custom_label_0") or "").strip()
        cat1 = (product.findtext("custom_label_1") or "").strip().replace("&amp;", "&")

        if not sku or not feed_title:
            continue

        # Offizielle Bezeichnung aus Excel bevorzugen
        if sku in excel_names:
            name = excel_names[sku]
            matched += 1
        else:
            name = feed_title
            fallback += 1

        articles.append({
            "sku": sku,
            "name": name,
            "category": cat0 if cat0 else "Sonstiges",
            "subcategory": cat1 if cat1 else "",
            "availability": availability,
            "price": price,
            "salePrice": sale_price if sale_price else None,
            "url": link,
            "gtin": gtin,
        })

    articles.sort(key=lambda x: (x["category"], x.get("subcategory", ""), x["name"]))
    print(f"  - {matched} mit offizieller Bezeichnung aus Excel")
    print(f"  - {fallback} mit Feed-Titel (nicht in Excel)")
    return articles


def generate_json(articles):
    """articles.json generieren."""
    data = {
        "meta": {
            "lastUpdated": str(date.today()),
            "totalArticles": len(articles),
            "source": "Shopware Feed (Kategorien) + Excel (offizielle Bezeichnungen)",
            "feedUrl": FEED_URL,
        },
        "categoryKeywords": CATEGORY_KEYWORDS,
        "articles": articles,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\narticles.json aktualisiert: {len(articles)} Artikel")


def print_summary(articles):
    """Zusammenfassung ausgeben."""
    from collections import Counter

    cats = Counter(a["category"] for a in articles)
    print("\nKategorien:")
    for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")
    print(f"\nGesamt: {len(articles)} Artikel")


if __name__ == "__main__":
    excel_names = load_excel_names()
    download_feed()
    articles = parse_feed(excel_names)
    generate_json(articles)
    print_summary(articles)
    print("\nFertig! Die Artikeldaten sind jetzt aktuell.")
